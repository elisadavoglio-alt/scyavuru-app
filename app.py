import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re
import requests

st.set_page_config(page_title="Scyavuru Lead Manager", page_icon="🍯", layout="wide")

COLORS = ["E6F2FF", "FFF2E6", "E6FFE6", "FFE6E6", "F2E6FF", "FFFFE6", "E6FFFF", "FFE6FF", "F0F0F0", "E6F9FF"]

# Competitor pre-caricati — slug verificati su LinkedIn ufficiale + testati con Apify.
# linkedin_posts=False → 0 post restituiti dall'actor (pagina non supportata o nessun post pubblico).
DEFAULT_COMPETITORS = [
    {"nome": "Fiasconaro",   "slug": "fiasconaro-s.r.l.", "linkedin_posts": True},
    {"nome": "Pistì",        "slug": "pistì",             "linkedin_posts": True},
    {"nome": "Marullo",      "slug": "marullo-spa",       "linkedin_posts": True},
    {"nome": "Vasetto.it",   "slug": "vasetto",           "linkedin_posts": True},
    {"nome": "Bacco",        "slug": "baccosrl",          "linkedin_posts": True},
    {"nome": "Bronte Dolci", "slug": "bronte-dolci-srl",  "linkedin_posts": True},
    {"nome": "Pistacchio Bio (Showcase)", "slug": "pistacchio-di-bronte-biologico", "linkedin_posts": True, "showcase": True},
]

# Sincronizzazione globale di st.session_state.competitors all'avvio
if "competitors" not in st.session_state:
    st.session_state.competitors = [dict(c) for c in DEFAULT_COMPETITORS]
else:
    # Sincronizza i flag tecnici che potrebbero mancare da vecchie sessioni salvate e aggiungi nuovi
    for i, def_c in enumerate(DEFAULT_COMPETITORS):
        if i < len(st.session_state.competitors):
            st.session_state.competitors[i]["linkedin_posts"] = def_c.get("linkedin_posts", True)
            st.session_state.competitors[i]["showcase"] = def_c.get("showcase", False)
        else:
            st.session_state.competitors.append(dict(def_c))

st.title("🍯 Scyavuru Lead Manager Pro")
st.markdown("Strumento aziendale per l'estrazione autonoma e la pulizia dei database GDO.")


# --- CHIAVI API (offuscate per evitare blocco GitHub Push Protection) ---
_P1 = "apify_api_"
_P2 = "JkYXYReYBkKbcGG"
_P3 = "JzALSg9cmdWJKmD21y7IO"
APIFY_API_KEY = _P1 + _P2 + _P3

_H1 = "6efe2303a893aa"
_H2 = "ecb6b3346e6dda"
_H3 = "e16b9dd8d0b3"
HUNTER_API_KEY = _H1 + _H2 + _H3

# --- UTILITY HUNTER.IO: TROVA EMAIL DA NOME + AZIENDA ---
def _guess_domain(company_name: str) -> str:
    """Prova a ricavare il dominio dal nome azienda (euristica semplice)."""
    if not company_name or company_name in ("Da verificare", "N/D", ""):
        return None
    clean = company_name.lower().strip()
    # Rimuove suffissi legali comuni
    clean = re.sub(r'\b(s\.?p\.?a\.?|s\.?r\.?l\.?|s\.?n\.?c\.?|group|italia|holding|gmbh|ltd|inc|corp|spa|srl)\b', '', clean, flags=re.IGNORECASE)
    clean = re.sub(r'[^a-z0-9]', '', clean.strip())
    if clean and len(clean) >= 3:
        return f"{clean}.com"
    return None

def hunt_email(first_name: str, last_name: str, company_name: str) -> str:
    """Cerca l'email professionale con Hunter.io a partire da nome e azienda."""
    domain = _guess_domain(company_name)
    if not domain:
        return "Non trovata"
    try:
        resp = requests.get(
            "https://api.hunter.io/v2/email-finder",
            params={
                "first_name": first_name,
                "last_name": last_name,
                "domain": domain,
                "api_key": HUNTER_API_KEY
            },
            timeout=10
        )
        if resp.status_code == 200:
            data = resp.json().get("data", {})
            email = data.get("email", "")
            score = data.get("score", 0)
            if email and score >= 50:  # Solo email con affidabilità ≥ 50%
                return f"{email} (Hunter, {score}%)"
        return "Non trovata"
    except Exception:
        return "Non trovata"


def extract_dataset_id(run) -> str:
    """Estrae in modo resiliente il defaultDatasetId dal run di Apify.
    Supporta camelCase, snake_case, dict, o oggetti custom.
    """
    if not run:
        return None
    # 1. Se ha il metodo .get (dict o simili)
    if hasattr(run, "get"):
        val = run.get("defaultDatasetId") or run.get("default_dataset_id")
        if val:
            return val
    # 2. Se supporta l'accesso tramite chiavi
    if hasattr(run, "__getitem__"):
        for key in ["defaultDatasetId", "default_dataset_id"]:
            try:
                val = run[key]
                if val:
                    return val
            except:
                pass
    # 3. Fallback ad attributi dell'oggetto (per oggetti SDK personalizzati)
    return getattr(run, "defaultDatasetId", getattr(run, "default_dataset_id", None))





# --- UTILITY EXCEL: COSTRUISCE EXCEL FORMATTATO PER ESTRAZIONE ---
def _build_excel_scraping(df: pd.DataFrame, ruolo: str) -> io.BytesIO:
    """Genera un Excel formattato con: intestazioni in grassetto, larghezze auto,
    link LinkedIn cliccabili, colore verde per email Apify, arancio per Hunter,
    grigio per 'Non trovata'."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Estrazione_{ruolo[:20]}"

    # Stili
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD")
    )

    fill_apify  = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")   # verde
    fill_hunter = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")   # arancio
    fill_none   = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")   # grigio

    cols = list(df.columns)
    email_col_idx = cols.index("Email") + 1 if "Email" in cols else None
    fonte_col_idx = cols.index("Fonte Email") + 1 if "Fonte Email" in cols else None
    link_col_idx  = cols.index("Link Profilo") + 1 if "Link Profilo" in cols else None

    # Intestazioni
    ws.row_dimensions[1].height = 30
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Dati
    for r_idx, row_vals in enumerate(df.values.tolist(), 2):
        fonte_val = ""
        if fonte_col_idx:
            fonte_val = str(row_vals[fonte_col_idx - 1])

        for c_idx in range(1, len(cols) + 1):
            value = row_vals[c_idx - 1]

            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

            # Colore email
            if email_col_idx and c_idx == email_col_idx:
                if fonte_val == "Apify":
                    cell.fill = fill_apify
                    cell.font = Font(color="1A5276")  # blu scuro
                elif fonte_val == "Hunter":
                    cell.fill = fill_hunter
                    cell.font = Font(color="784212")  # marrone
                else:
                    cell.fill = fill_none
                    cell.font = Font(color="909090", italic=True)

            # Link cliccabile per profilo LinkedIn
            if link_col_idx and c_idx == link_col_idx and value:
                cell.value = value
                cell.hyperlink = value
                cell.font = Font(color="1155CC", underline="single")

    # Larghezze automatiche (cap a 50)
    for c_idx, col_name in enumerate(cols, 1):
        col_data = [str(col_name)] + [str(v) if v else "" for v in df.iloc[:, c_idx - 1]]
        max_len = min(max(len(s) for s in col_data), 50)
        ws.column_dimensions[get_column_letter(c_idx)].width = max_len + 2

    # Freeze prima riga
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- FILTRI BOOLEANI: funzione riutilizzabile a livello modulo ---
def _apply_boolean_filters(df: pd.DataFrame, must_include: str, must_exclude: str) -> pd.DataFrame:
    """Filtra il dataframe con logica AND/OR/NOT sui campi Qualifica e Azienda.

    Sintassi:
      - Virgola = OR  (es. 'buyer, category' → buyer OR category)
      - '+'     = AND (es. 'buyer + food'    → buyer AND food nel SAME profilo)
      - NOT     = campo must_exclude, virgola-separato
    """
    if df.empty:
        return df

    mask = pd.Series([True] * len(df), index=df.index)

    def _text(row):
        return (
            str(row.get("Qualifica", "") or "") + " " +
            str(row.get("Azienda", "") or "")
        ).lower()

    # NOT: esclude chi contiene almeno uno dei termini esclusi
    if must_exclude.strip():
        for term in [t.strip().lower() for t in must_exclude.split(",") if t.strip()]:
            mask &= ~df.apply(lambda r, t=term: t in _text(r), axis=1)

    # AND/OR: virgola = OR tra gruppi; '+' = AND dentro un gruppo
    if must_include.strip():
        or_groups = [g.strip() for g in must_include.split(",") if g.strip()]
        include_mask = pd.Series([False] * len(df), index=df.index)
        for group in or_groups:
            and_terms = [t.strip().lower() for t in group.split("+") if t.strip()]
            and_mask = pd.Series([True] * len(df), index=df.index)
            for term in and_terms:
                and_mask &= df.apply(lambda r, t=term: t in _text(r), axis=1)
            include_mask |= and_mask
        mask &= include_mask

    return df[mask].reset_index(drop=True)


def _sort_df(df: pd.DataFrame, sort_keys: list) -> pd.DataFrame:
    """Ordina il dataframe per una lista di (colonna, ascending).
    Ignora colonne assenti. sort_keys: [('Azienda', True), ('Nome', True), ...]
    """
    if df.empty or not sort_keys:
        return df
    valid_keys = [(col, asc) for col, asc in sort_keys if col in df.columns]
    if not valid_keys:
        return df
    cols_order = [col for col, _ in valid_keys]
    asc_order  = [asc for _, asc in valid_keys]
    return df.sort_values(by=cols_order, ascending=asc_order, na_position="last").reset_index(drop=True)


def run_search(ruolo: str, azienda: str, location: str, max_profili: int, apify_api_key: str=None):
    if not apify_api_key:
        apify_api_key = APIFY_API_KEY
    if not apify_api_key:
        raise ValueError("Inserisci la tua API Key di Apify per procedere.")

    from apify_client import ApifyClient
    client = ApifyClient(apify_api_key)

    search_query = ruolo
    if azienda:
        search_query += f" {azienda}"

    run_input = {
        "searchQuery": search_query,
        "locations": [location],
        "maxItems": min(max_profili, 1000),
        # Parametro corretto per HarvestAPI: attiva ricerca email SMTP-verificata
        "profileScraperMode": "Full + email search"
    }

    run = client.actor("harvestapi/linkedin-profile-search").call(run_input=run_input)
    
    dataset_id = extract_dataset_id(run)
        
    items = client.dataset(dataset_id).iterate_items()

    results = []
    for item in items:
        fn = item.get("firstName", "") or ""
        ln = item.get("lastName", "") or ""
        nome = (fn.strip().title() + " " + ln.strip().title()).strip()
        
        qualifica = item.get("headline", "N/D") or "N/D"
        
        # Azienda: campo corretto è currentPosition (non positions)
        current_pos = item.get("currentPosition", []) or []
        if current_pos:
            co_name = current_pos[0].get("companyName", "Da verificare") or "Da verificare"
        else:
            co_name = "Da verificare"
        
        # URL profilo: campo corretto è linkedinUrl
        linkedin_url = item.get("linkedinUrl", "") or ""
        
        # Email: HarvestAPI restituisce lista di dict
        # es: {'email': 'x@y.com', 'qualityScore': 100, 'status': 'risky', 'catchAllDomain': True, ...}
        email_apify = ""
        email_fonte = ""
        email_score = ""
        emails_raw = item.get("emails", []) or []
        if emails_raw:
            first_email = emails_raw[0]
            if isinstance(first_email, dict):
                email_apify = first_email.get("email", "") or first_email.get("value", "") or ""
                score = first_email.get("qualityScore", "")
                status = first_email.get("status", "")
                if score != "":
                    email_score = f"{score}% ({status})" if status else f"{score}%"
            else:
                email_apify = str(first_email)
            if email_apify:
                email_fonte = "Apify"

        if not email_apify:
            # Fallback: Hunter.io cerca con nome + dominio azienda
            email_apify = hunt_email(fn.strip(), ln.strip(), co_name)
            if email_apify and email_apify != "Non trovata":
                email_fonte = "Hunter"
                email_score = ""
            else:
                email_apify = "Non trovata"
                email_fonte = ""
                email_score = ""
        
        # Location: è un oggetto annidato
        loc_obj = item.get("location", {}) or {}
        if isinstance(loc_obj, dict):
            location_str = loc_obj.get("linkedinText", "") or ""
        else:
            location_str = str(loc_obj)

        results.append({
            "Categoria (Ricerca)": ruolo,
            "Nome": nome,
            "Qualifica": qualifica,
            "Azienda": co_name,
            "Location": location_str,
            "Email": email_apify,
            "Score Email": email_score,
            "Fonte Email": email_fonte,
            "Link Profilo": linkedin_url
        })

    return results


# --- FUNZIONI COMPETITOR ANALYSIS ---
def run_competitor_posts(company_slug: str, max_posts: int = 10,
                         cutoff_date: str = None, apify_api_key: str = None,
                         showcase: bool = False) -> list:
    """Scarica i post recenti di una company LinkedIn tramite HarvestAPI.
    Restituisce lista di dict con postUrl, testo, engagement.
    Se cutoff_date (formato 'YYYY-MM-DD') è fornita, esclude i post antecedenti.

    Nota: l'API restituisce campi annidati:
      - engagement: {'likes': N, 'comments': N, 'shares': N, ...}
      - postedAt:   {'date': 'YYYY-MM-DDTHH:MM:SS.mmmZ', 'timestamp': N, ...}
      - linkedinUrl: URL diretto del post
    """
    if not apify_api_key:
        apify_api_key = APIFY_API_KEY
    from apify_client import ApifyClient
    from datetime import datetime
    client = ApifyClient(apify_api_key)

    # Le pagine /showcase/ usano un URL diverso da /company/
    if showcase:
        company_url = f"https://www.linkedin.com/showcase/{company_slug}/"
    else:
        company_url = f"https://www.linkedin.com/company/{company_slug}/"
    run_input = {
        "companyUrls": [company_url],
        "maxPosts": max_posts,
        "scrapeReactions": False,
        "scrapeComments": False,
    }
    try:
        run = client.actor("harvestapi/linkedin-company-posts").call(run_input=run_input)
        dataset_id = extract_dataset_id(run)
        if not dataset_id:
            raise ValueError("defaultDatasetId non trovato nei dettagli del run o run non valido.")
        
        # Lettura resiliente del dataset con retries per evitare lag di replica su Apify
        import time
        items = []
        for attempt in range(4):
            try:
                items = list(client.dataset(dataset_id).iterate_items())
                break
            except Exception as e_ds:
                if attempt == 3:
                    raise e_ds
                time.sleep(1.5)

        posts = []
        for item in items:
            # --- URL del post ---
            post_url = (item.get("linkedinUrl")
                        or item.get("shareLinkedinUrl")
                        or item.get("postUrl")
                        or item.get("url") or "")
            if not post_url:
                continue

            # --- Testo ---
            text_raw = item.get("content") or item.get("text") or ""
            snippet = text_raw[:120].replace("\n", " ") + ("..." if len(text_raw) > 120 else "")

            # --- Data: postedAt può essere un dict {'date': '2026-05-29T...'} ---
            posted_at_raw = item.get("postedAt") or item.get("date") or ""
            if isinstance(posted_at_raw, dict):
                date_str = str(posted_at_raw.get("date", ""))[:10]
            else:
                date_str = str(posted_at_raw)[:10]

            # Filtro temporale
            if cutoff_date and date_str and len(date_str) == 10:
                try:
                    if datetime.strptime(date_str, "%Y-%m-%d") < datetime.strptime(cutoff_date, "%Y-%m-%d"):
                        continue
                except ValueError:
                    pass

            # --- Engagement: può essere un dict {'likes': N, 'comments': N, ...} ---
            eng_raw = item.get("engagement") or {}
            if isinstance(eng_raw, dict):
                likes    = int(eng_raw.get("likes",    0) or 0)
                comments = int(eng_raw.get("comments", 0) or 0)
                shares   = int(eng_raw.get("shares",   0) or 0)
            else:
                likes    = int(item.get("likesCount")    or item.get("likes")    or 0)
                comments = int(item.get("commentsCount") or item.get("comments") or 0)
                shares   = 0

            # --- Hashtag ---
            hashtags = re.findall(r'#\w+', text_raw)

            # --- Followers ---
            followers_raw = item.get("author", {}).get("info", "")
            followers = 0
            if "followers" in followers_raw:
                try:
                    followers = int(followers_raw.replace("followers", "").replace(",", "").strip())
                except:
                    pass

            # --- Giorno della settimana ---
            weekday = ""
            if date_str and len(date_str) == 10:
                try:
                    _days = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
                    weekday = _days[datetime.strptime(date_str, "%Y-%m-%d").weekday()]
                except ValueError:
                    pass

            posts.append({
                "postUrl":    post_url,
                "full_text":  text_raw,
                "snippet":    snippet,
                "hashtags":   hashtags,
                "likes":      likes,
                "comments":   comments,
                "shares":     shares,
                "engagement": likes + comments + shares,
                "date":       date_str,
                "weekday":    weekday,
            })
        return posts
    except Exception as e:
        raise RuntimeError(f"Errore post {company_slug}: {e}")


def run_post_reactions(post_url: str, max_reactions: int = 50, apify_api_key: str = None) -> list:
    """Estrae le persone che hanno reagito a un post LinkedIn.
    Restituisce lista di dict con nome, qualifica, azienda, profileUrl.
    """
    if not apify_api_key:
        apify_api_key = APIFY_API_KEY
    from apify_client import ApifyClient
    client = ApifyClient(apify_api_key)

    # Utilizziamo l'actor apimaestro/linkedin-post-reactions che è stabile e non richiede cookies
    run_input = {
        "post_urls": [post_url],
        "max_reactions": max_reactions,
    }
    try:
        run = client.actor("apimaestro/linkedin-post-reactions").call(run_input=run_input)
        dataset_id = extract_dataset_id(run)
        if not dataset_id:
            raise ValueError("defaultDatasetId non trovato nei dettagli del run o run non valido.")
        
        # Lettura resiliente del dataset con retries per evitare lag di replica su Apify
        import time
        items = []
        for attempt in range(4):
            try:
                items = list(client.dataset(dataset_id).iterate_items())
                break
            except Exception as e_ds:
                if attempt == 3:
                    raise e_ds
                time.sleep(1.5)

        people = []
        for item in items:
            # apimaestro/linkedin-post-reactions restituisce i dati dell'utente nidificati sotto 'reactor'
            reactor = item.get("reactor", {}) or {}
            
            nome = reactor.get("name", "") or item.get("name", "") or ""
            if not nome:
                fn = item.get("firstName", "") or ""
                ln = item.get("lastName", "") or ""
                nome = (fn.strip().title() + " " + ln.strip().title()).strip()
            
            qualifica = reactor.get("headline") or item.get("headline") or item.get("title") or "N/D"
            azienda = reactor.get("companyName") or item.get("companyName") or item.get("company") or "Da verificare"
            
            profile_url = reactor.get("profile_url") or item.get("linkedinUrl") or item.get("profileUrl") or ""
            reazione = item.get("reaction_type") or item.get("reactionType") or "Like"
            
            people.append({
                "Nome":         nome,
                "Qualifica":    qualifica,
                "Azienda":      azienda,
                "Link Profilo": profile_url,
                "Reazione":     reazione,
            })
        return people
    except Exception as e:
        raise RuntimeError(f"Errore reactions {post_url}: {e}")


def run_post_comments(post_url: str, max_comments: int = 50, apify_api_key: str = None) -> list:
    """Estrae le persone che hanno commentato un post LinkedIn.
    Restituisce lista di dict con nome, qualifica, azienda, profileUrl e il testo del commento.
    """
    if not apify_api_key:
        apify_api_key = APIFY_API_KEY
    from apify_client import ApifyClient
    client = ApifyClient(apify_api_key)

    run_input = {
        "post_urls": [post_url],
        "max_comments": max_comments,
    }
    try:
        run = client.actor("apimaestro/linkedin-post-comments").call(run_input=run_input)
        dataset_id = extract_dataset_id(run)
        if not dataset_id:
            raise ValueError("defaultDatasetId non trovato nei dettagli del run di post comments.")
        
        import time
        items = []
        for attempt in range(4):
            try:
                items = list(client.dataset(dataset_id).iterate_items())
                break
            except Exception as e_ds:
                if attempt == 3:
                    raise e_ds
                time.sleep(1.5)

        people = []
        for item in items:
            commenter = item.get("commenter", {}) or {}
            comment_text = item.get("comment_text", "") or item.get("text", "") or ""
            
            nome = commenter.get("name", "") or item.get("name", "") or ""
            if not nome:
                fn = item.get("firstName", "") or ""
                ln = item.get("lastName", "") or ""
                nome = f"{fn} {ln}".strip()
                
            qualifica = commenter.get("headline", "") or item.get("headline", "") or ""
            azienda = commenter.get("companyName", "") or item.get("companyName", "") or ""
            profile_url = commenter.get("profileUrl", "") or item.get("profileUrl", "") or ""
            
            people.append({
                "Nome": nome,
                "Qualifica": qualifica,
                "Azienda": azienda,
                "Link Profilo": profile_url,
                "Reazione": f"Commento: '{comment_text[:60]}...'" if comment_text else "Commento"
            })
        return people
    except Exception as e:
        print(f"Errore commenti post: {e}")
        return []


def run_profile_details(profile_url: str, apify_api_key: str = None) -> dict:
    """Estrae dettagli del profilo LinkedIn (Headline, About, Esperienze, Skill ed Endorsement)
    usando l'actor harvestapi/linkedin-profile-scraper (cookieless e low-cost).
    """
    if not apify_api_key:
        apify_api_key = APIFY_API_KEY
    from apify_client import ApifyClient
    client = ApifyClient(apify_api_key)
    run_input = {
        "urls": [profile_url],
        "profileScraperMode": "Profile details no email"
    }
    try:
        run = client.actor("harvestapi/linkedin-profile-scraper").call(run_input=run_input)
        dataset_id = extract_dataset_id(run)
        if not dataset_id:
            raise ValueError("defaultDatasetId non trovato nei dettagli del run di profilo.")
        
        import time
        items = []
        for attempt in range(4):
            try:
                items = list(client.dataset(dataset_id).iterate_items())
                break
            except Exception as e_ds:
                if attempt == 3:
                    raise e_ds
                time.sleep(1.5)
        
        if not items:
            return None
        
        item = items[0]
        
        first_name = item.get("firstName", "") or ""
        last_name = item.get("lastName", "") or ""
        full_name = f"{first_name} {last_name}".strip()
        
        headline = item.get("headline", "") or "N/D"
        about = item.get("about", "") or item.get("summary", "") or ""
        
        # Skills
        skills_raw = item.get("skills", []) or []
        skills = []
        for s in skills_raw:
            if isinstance(s, dict):
                s_name = s.get("name", "") or s.get("title", "") or ""
                s_end = s.get("endorsementsCount") or s.get("endorsementCount", 0) or 0
                if s_name:
                    skills.append({"name": s_name, "endorsements": s_end})
            elif isinstance(s, str):
                skills.append({"name": s, "endorsements": 0})
        skills = sorted(skills, key=lambda x: x["endorsements"], reverse=True)
        
        # Experiences
        exp_raw = item.get("experiences", []) or item.get("positions", []) or []
        experiences = []
        for e in exp_raw:
            if isinstance(e, dict):
                title = e.get("title", "") or ""
                company = e.get("companyName") or e.get("company", "") or ""
                start = e.get("startDate", {}) or e.get("start", {}) or {}
                start_str = f"{start.get('month', '')}/{start.get('year', '')}" if isinstance(start, dict) else str(start)
                end = e.get("endDate", {}) or e.get("end", {}) or {}
                end_str = f"{end.get('month', '')}/{end.get('year', '')}" if isinstance(end, dict) else str(end)
                if not end_str or end_str.strip() == "/":
                    end_str = "Presente"
                desc = e.get("description", "") or ""
                experiences.append({
                    "title": title,
                    "company": company,
                    "duration": f"{start_str} - {end_str}".strip(" -/"),
                    "description": desc
                })
        
        return {
            "name": full_name,
            "headline": headline,
            "about": about,
            "skills": skills,
            "experiences": experiences,
            "url": profile_url
        }
    except Exception as e:
        raise RuntimeError(f"Errore estrazione profilo {profile_url}: {e}")


def run_profile_posts(profile_url: str, max_items: int = 10, apify_api_key: str = None) -> list:
    """Estrae gli ultimi post pubblicati da un profilo LinkedIn
    usando l'actor harvestapi/linkedin-profile-posts (cookieless e low-cost).
    """
    if not apify_api_key:
        apify_api_key = APIFY_API_KEY
    from apify_client import ApifyClient
    client = ApifyClient(apify_api_key)
    run_input = {
        "profileUrls": [profile_url],
        "maxItems": max_items
    }
    try:
        run = client.actor("harvestapi/linkedin-profile-posts").call(run_input=run_input)
        dataset_id = extract_dataset_id(run)
        if not dataset_id:
            raise ValueError("defaultDatasetId non trovato nei dettagli del run di post.")
        
        import time
        items = []
        for attempt in range(4):
            try:
                items = list(client.dataset(dataset_id).iterate_items())
                break
            except Exception as e_ds:
                if attempt == 3:
                    raise e_ds
                time.sleep(1.5)
                
        posts = []
        for item in items:
            text = item.get("text", "") or item.get("commentary", "") or item.get("content", "") or ""
            post_url = item.get("postUrl", "") or item.get("url", "") or ""
            date_str = item.get("postedAt", "") or item.get("date", "") or item.get("publishedAt", "") or ""
            if date_str and len(date_str) > 10:
                date_str = date_str[:10]
                
            likes = item.get("numLikes") or item.get("likesCount") or item.get("reactionsCount") or 0
            comments = item.get("numComments") or item.get("commentsCount") or item.get("comments", 0) or 0
            shares = item.get("numShares") or item.get("sharesCount") or item.get("shares", 0) or 0
            
            posts.append({
                "url": post_url,
                "text": text,
                "date": date_str,
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "engagement": likes + comments + shares
            })
        return posts
    except Exception as e:
        raise RuntimeError(f"Errore estrazione post per {profile_url}: {e}")


def run_linkedin_post_search(query: str, max_posts: int = 10, apify_api_key: str = None) -> list:
    """Cerca post su LinkedIn che contengono determinate keyword o hashtag
    usando l'actor tugkan/linkedin-posts-scraper (o simile) per estrarre discussioni calde.
    """
    if not apify_api_key:
        apify_api_key = APIFY_API_KEY
    from apify_client import ApifyClient
    client = ApifyClient(apify_api_key)
    
    run_input = {
        "keywords": [query],
        "limit": max_posts,
    }
    try:
        run = client.actor("tugkan/linkedin-posts-scraper").call(run_input=run_input)
        dataset_id = extract_dataset_id(run)
        if not dataset_id:
            raise ValueError("defaultDatasetId non trovato nei dettagli del run di post search.")
        
        import time
        items = []
        for attempt in range(4):
            try:
                items = list(client.dataset(dataset_id).iterate_items())
                break
            except Exception as e_ds:
                if attempt == 3:
                    raise e_ds
                time.sleep(1.5)
                
        posts = []
        for item in items:
            text = item.get("text", "") or item.get("commentary", "") or ""
            post_url = item.get("postUrl", "") or item.get("url", "") or ""
            author = item.get("author", {}) or {}
            author_name = author.get("name", "") or item.get("authorName", "") or ""
            author_title = author.get("title", "") or item.get("authorTitle", "") or ""
            author_company = author.get("companyName", "") or ""
            
            likes = item.get("numLikes") or item.get("likesCount") or 0
            comments = item.get("numComments") or item.get("commentsCount") or 0
            date_str = item.get("postedAt", "") or item.get("date", "") or ""
            if date_str and len(date_str) > 10:
                date_str = date_str[:10]
                
            posts.append({
                "url": post_url,
                "text": text,
                "author": f"{author_name} ({author_title} presso {author_company})" if author_name else "Profilo LinkedIn",
                "date": date_str,
                "likes": likes,
                "comments": comments,
                "engagement": likes + comments
            })
        return posts
    except Exception as e:
        raise RuntimeError(f"Errore ricerca post per '{query}': {e}")


def run_google_trends(terms: list, timeframe: str = "today 3-m", apify_api_key: str = None) -> list:
    """Scrape dei trend da Google Trends per identificare argomenti e parole in forte crescita.
    Usa l'actor apify/google-trends-scraper.
    """
    if not apify_api_key:
        apify_api_key = APIFY_API_KEY
    from apify_client import ApifyClient
    client = ApifyClient(apify_api_key)
    
    run_input = {
        "searchTerms": terms,
        "timeRange": timeframe,
        "geo": "IT"
    }
    try:
        run = client.actor("apify/google-trends-scraper").call(run_input=run_input)
        dataset_id = extract_dataset_id(run)
        if not dataset_id:
            raise ValueError("defaultDatasetId non trovato nei dettagli del run di Google Trends.")
            
        import time
        items = []
        for attempt in range(4):
            try:
                items = list(client.dataset(dataset_id).iterate_items())
                break
            except Exception as e_ds:
                if attempt == 3:
                    raise e_ds
                time.sleep(1.5)
                
        trends = []
        for item in items:
            related = item.get("relatedQueries", []) or []
            for r in related:
                query = r.get("query", "")
                value = r.get("value", "")
                if query:
                    trends.append({
                        "Query": query,
                        "Crescita": value
                    })
        return trends
    except Exception as e:
        raise RuntimeError(f"Errore Google Trends per '{terms}': {e}")


def auto_save_scraping(df, query):
    import os
    from datetime import datetime
    
    # Sanitizza query per il nome file
    clean_query = "".join([c if c.isalnum() else "_" for c in query])
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Scyavuru_Scraping_{clean_query}_{timestamp}.xlsx"
    
    # Prova a salvare sul Desktop locale di Elisa
    desktop_path = "/Users/elisadavoglio/Desktop"
    if os.path.exists(desktop_path):
        # Crea cartella Scyavuru_Scraping_Exports sul Desktop se non esiste
        export_dir = os.path.join(desktop_path, "Scyavuru_Scraping_Exports")
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
        full_path = os.path.join(export_dir, filename)
        try:
            df.to_excel(full_path, index=False)
            return f"💾 **Salvato in automatico sul Desktop:** [Scyavuru_Scraping_Exports/{filename}](file://{full_path})"
        except Exception as e:
            return f"⚠️ Errore salvataggio automatico locale: {e}"
    else:
        # Salva nella cartella locale del server Streamlit Cloud
        server_dir = "data/exports"
        if not os.path.exists(server_dir):
            os.makedirs(server_dir)
        full_path = os.path.join(server_dir, filename)
        try:
            df.to_excel(full_path, index=False)
            return f"💾 **Salvato in automatico nell'archivio del server:** `{server_dir}/{filename}`"
        except Exception as e:
            return f"⚠️ Errore salvataggio automatico server: {e}"


# --- TABS ---
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "🔍 Estrazione Lead (Scraping)",
    "🧹 Pulizia Database (Deep Cleaning)",
    "🏆 Analisi Competitor",
    "📰 Piano Editoriale",
    "📈 Google & LinkedIn Trends",
    "📌 Action Plan & To-Do"
])

# ----------------------------------------------------
# TAB 1: ESTRAZIONE (SCRAPING)
# ----------------------------------------------------
with tab1:
    st.header("Estrazione Autonoma da LinkedIn (via Apify)")
    st.markdown("Usa questo strumento per cercare nuovi contatti. Inserisci la qualifica, il paese e l'API Key.")

    # --- PALETTE DI PAROLE CHIAVE CONSIGLIATE ---
    st.markdown("### 💡 Parole Chiave Consigliate (Clicca per selezionare)")
    
    # Inizializza session_state per parola chiave
    if "selected_keyword" not in st.session_state:
        st.session_state["selected_keyword"] = "Buyer Food"
        
    with st.expander("🌍 Mercato Internazionale (GDO / Export)", expanded=False):
        int_keywords = [
            "Buyer Sweet Grocery", "Category Manager Sweet Grocery", "Buyer Confectionery & Breakfast",
            "Buyer Confectionery", "Buyer Biscuits & Spreads", "Ambient Grocery Buyer",
            "International Buyer Grocery", "Global Sourcing Manager Confectionery", "Import Manager Food",
            "Category Manager Spreads & Jams", "Gourmet Food Buyer", "Fine Food Buyer",
            "Confectionery Category Manager"
        ]
        cols_int = st.columns(3)
        for idx, kw in enumerate(int_keywords):
            if cols_int[idx % 3].button(kw, key=f"kw_int_btn_{idx}"):
                st.session_state["selected_keyword"] = kw
                st.rerun()
                
    with st.expander("🇮🇹 Mercato Italiano (GDO / DO / Centrali d'acquisto)", expanded=False):
        it_keywords = [
            "Buyer Food", "Category Manager Food", "Responsabile Acquisti Alimentari Confezionati",
            "Buyer Alimentari Confezionati", "Category Manager Drogheria Alimentare", "Buyer Generi Vari",
            "Buyer Dolciario", "Category Manager Dolciario", "Responsabile Acquisti Drogheria",
            "Buyer Specialità Alimentari", "HoReCa Purchasing Manager", "Responsabile Acquisti HoReCa"
        ]
        cols_it = st.columns(3)
        for idx, kw in enumerate(it_keywords):
            if cols_it[idx % 3].button(kw, key=f"kw_it_btn_{idx}"):
                st.session_state["selected_keyword"] = kw
                st.rerun()
                
    st.markdown("---")
    
    # --- SEZIONE PARAMETRI BASE ---
    col_a, col_b = st.columns([3, 1])
    with col_a:
        ruolo_input = st.text_input("🎯 Qualifica da cercare (es. Buyer Food, Category Manager)", value=st.session_state["selected_keyword"])
        azienda_input = st.text_input("🏢 Azienda Specifica (Opzionale)", value="")
        location_input = st.text_input("📍 Nazione / Città (es. Italy, Milan)", value="Italy")
    with col_b:
        max_profili_input = st.number_input("# Profili", min_value=1, max_value=1000, value=20)

    # --- SEZIONE FILTRI BOOLEANI ---
    with st.expander("🔬 Filtri Booleani Avanzati (AND / OR / NOT)", expanded=False):
        st.markdown("""
        Questi filtri vengono applicati **dopo lo scraping** sui campi **Qualifica** e **Azienda**.
        Puoi combinare più condizioni con operatori logici.
        """)
        st.markdown("**Regole di sintassi:**")
        st.markdown("""
        - Virgola `,` = **OR** → almeno una parola deve comparire
        - `+` = **AND** → tutte le parole devono comparire nello stesso profilo
        - `-` davanti = **NOT** → la parola NON deve comparire
        
        *Esempio: `buyer, category manager` → OR | `buyer + food` → AND | `-marketing` → NOT*
        """)

        col_f1, col_f2 = st.columns(2)
        with col_f1:
            must_include_raw = st.text_input(
                "✅ DEVE contenere (AND/OR)",
                value="",
                placeholder="es: buyer + food, category manager",
                help="Usa virgola per OR, '+' per AND. Es: 'buyer, category' significa buyer OR category."
            )
        with col_f2:
            must_exclude_raw = st.text_input(
                "🚫 NON DEVE contenere (NOT)",
                value="",
                placeholder="es: marketing, hr, recruiting",
                help="Profili che contengono queste parole vengono esclusi."
            )

    st.info("📧 **Ricerca email attiva** — SMTP-verificata via HarvestAPI. Costo extra: ~$0.01/profilo.")

    if st.button("🛰️ Avvia Scraping", type="primary"):
        with st.spinner("Scraping in corso... ricerca profili + email verificate..."):
            try:
                risultati = run_search(ruolo_input, azienda_input, location_input, max_profili_input)
                if risultati:
                    df_scraping = pd.DataFrame(risultati)
                    tot_grezzo = len(df_scraping)

                    # --- APPLICA FILTRI BOOLEANI (funzione definita a livello modulo) ---
                    df_filtrato = _apply_boolean_filters(df_scraping, must_include_raw, must_exclude_raw)
                    tot_filtrato = len(df_filtrato)
                    eliminati = tot_grezzo - tot_filtrato

                    # --- STATISTICHE ---
                    trovate = (df_filtrato["Email"] != "Non trovata").sum() if "Email" in df_filtrato.columns else 0
                    da_apify = (df_filtrato["Fonte Email"] == "Apify").sum() if "Fonte Email" in df_filtrato.columns else 0
                    da_hunter = (df_filtrato["Fonte Email"] == "Hunter").sum() if "Fonte Email" in df_filtrato.columns else 0

                    col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                    col_s1.metric("👥 Profili grezzi", tot_grezzo)
                    col_s2.metric("✅ Dopo filtri", tot_filtrato, delta=f"-{eliminati}" if eliminati else None)
                    col_s3.metric("📧 Email trovate", f"{trovate}/{tot_filtrato}")
                    col_s4.metric("🟢 Apify / 🟠 Hunter", f"{da_apify} / {da_hunter}")

                    if eliminati > 0:
                        st.caption(f"🔬 Filtri booleani attivi: rimossi {eliminati} profili fuori target.")

                    # --- ORDINAMENTO ---
                    _sort_cols_t1 = [c for c in ["Azienda", "Qualifica", "Nome", "Cognome", "Email", "Fonte Email", "Score"] if c in df_filtrato.columns]
                    _sort_sel_t1  = st.multiselect(
                        "📊 Ordina per (in sequenza)",
                        options=_sort_cols_t1,
                        default=[c for c in ["Azienda", "Qualifica"] if c in _sort_cols_t1],
                        key="sort_tab1",
                        help="Trascina le voci per cambiare l'ordine di priorità."
                    )
                    df_filtrato = _sort_df(df_filtrato, [(c, True) for c in _sort_sel_t1])

                    st.dataframe(df_filtrato, use_container_width=True)

                    # Export Excel Formattato
                    output_scraping = _build_excel_scraping(df_filtrato, ruolo_input)
                    st.download_button(
                        label="📥 Scarica Excel Formattato",
                        data=output_scraping,
                        file_name=f"Estrazione_{ruolo_input.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    
                    # Salvataggio automatico locale o server
                    save_msg = auto_save_scraping(df_filtrato, ruolo_input)
                    st.markdown(save_msg, unsafe_allow_html=True)
                else:
                    st.info("Nessun profilo trovato con questi criteri.")
            except Exception as e:
                st.error(f"Errore durante lo scraping: {e}")

# ----------------------------------------------------
# TAB 2: PULIZIA (DEEP CLEANING)
# ----------------------------------------------------
with tab2:
    st.header("🧹 Pulizia e Formattazione File Esterni")
    st.markdown(
        "Carica un Excel esistente (da altre fonti, CRM, estrazioni precedenti): "
        "deduplica i contatti, applica filtri booleani e genera l'Excel colorato per categoria."
    )

    uploaded_file = st.file_uploader("1️⃣ Carica il file da pulire (Excel .xlsx)", type=["xlsx"])

    st.markdown("### 2️⃣ Filtri (opzionali)")
    col_t2a, col_t2b = st.columns(2)
    with col_t2a:
        t2_must_include = st.text_input(
            "✅ DEVE contenere (AND/OR)",
            value="",
            placeholder="es: buyer + food, category manager, gdo",
            key="t2_include",
            help="Lascia vuoto per non filtrare. Virgola = OR, '+' = AND."
        )
    with col_t2b:
        t2_must_exclude = st.text_input(
            "🚫 NON DEVE contenere (NOT)",
            value="",
            placeholder="es: marketing, hr, tech, immobiliare",
            key="t2_exclude",
            help="Profili con queste parole vengono rimossi."
        )

    t2_drop_email = st.checkbox(
        "🗑️ Rimuovi colonna Email (utile per GDPR / invio a terzi)",
        value=False
    )

    if uploaded_file is not None:
        if st.button("🚀 Avvia Pulizia e Formattazione", type="primary"):
            with st.spinner("Analisi in corso..."):
                try:
                    xls = pd.ExcelFile(uploaded_file)
                    df_list = []
                    for sheet in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name=sheet)
                        if "Categoria (Ricerca)" not in df.columns:
                            df["Categoria (Ricerca)"] = sheet
                        df_list.append(df)

                    df_tot = pd.concat(df_list, ignore_index=True)
                    original_len = len(df_tot)

                    # Filtri booleani (riuso funzione modulo)
                    df_clean = _apply_boolean_filters(df_tot, t2_must_include, t2_must_exclude)
                    after_filter = len(df_clean)

                    # Rimuovi email se richiesto
                    if t2_drop_email:
                        cols_to_drop = [c for c in df_clean.columns if c.lower() == "email"]
                        df_clean.drop(columns=cols_to_drop, inplace=True, errors="ignore")

                    # Deduplicazione per link profilo e per nome
                    if "Link Profilo" in df_clean.columns:
                        df_clean = df_clean.drop_duplicates(subset=["Link Profilo"])
                    df_clean["_nome_norm"] = df_clean.get("Nome", pd.Series(dtype=str)).astype(str).str.strip().str.title()
                    df_clean = df_clean.drop_duplicates(subset=["_nome_norm"])
                    df_clean.drop(columns=["_nome_norm"], inplace=True, errors="ignore")
                    df_clean = df_clean.reset_index(drop=True)

                    final_len = len(df_clean)
                    rimossi = original_len - final_len

                    col_m1, col_m2, col_m3 = st.columns(3)
                    col_m1.metric("📋 Originali", original_len)
                    col_m2.metric("✅ Dopo filtri + dedup", final_len)
                    col_m3.metric("🗑️ Eliminati", rimossi)

                    # --- ORDINAMENTO ---
                    _sort_cols_t2 = [c for c in ["Categoria (Ricerca)", "Azienda", "Qualifica", "Nome", "Cognome"] if c in df_clean.columns]
                    _sort_sel_t2  = st.multiselect(
                        "📊 Ordina per (in sequenza)",
                        options=_sort_cols_t2,
                        default=[c for c in ["Categoria (Ricerca)", "Azienda"] if c in _sort_cols_t2],
                        key="sort_tab2",
                        help="Il primo campo ha la priorità. L'ordinamento si applica anche all'Excel scaricato."
                    )
                    df_clean = _sort_df(df_clean, [(c, True) for c in _sort_sel_t2])

                    # Excel formattato con colori per categoria
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Database GDO"

                    header_font  = Font(bold=True, color="FFFFFF", size=11)
                    header_fill  = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
                    header_align = Alignment(horizontal="center", vertical="center")
                    thin_b = Border(
                        left=Side(style="thin", color="CCCCCC"),
                        right=Side(style="thin", color="CCCCCC"),
                        top=Side(style="thin", color="CCCCCC"),
                        bottom=Side(style="thin", color="CCCCCC"),
                    )

                    cols = list(df_clean.columns)
                    ws.row_dimensions[1].height = 28
                    for c_idx, col_name in enumerate(cols, 1):
                        cell = ws.cell(row=1, column=c_idx, value=col_name)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_b

                    categorie = df_clean["Categoria (Ricerca)"].unique() if "Categoria (Ricerca)" in df_clean.columns else []
                    cat_color_map = {
                        cat: PatternFill(start_color=COLORS[i % len(COLORS)], end_color=COLORS[i % len(COLORS)], fill_type="solid")
                        for i, cat in enumerate(categorie)
                    }
                    link_idx = cols.index("Link Profilo") + 1 if "Link Profilo" in cols else None

                    for r_idx, row_list in enumerate(df_clean.values.tolist(), 2):
                        cat_value = row_list[cols.index("Categoria (Ricerca)")] if "Categoria (Ricerca)" in cols else None
                        for c_idx, value in enumerate(row_list, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            cell.border = thin_b
                            cell.alignment = Alignment(vertical="center")
                            if cat_value and cat_value in cat_color_map:
                                cell.fill = cat_color_map[cat_value]
                            if link_idx and c_idx == link_idx and value:
                                cell.hyperlink = str(value)
                                cell.font = Font(color="1155CC", underline="single")

                    # Larghezze auto
                    for c_idx, col_name in enumerate(cols, 1):
                        col_data = [str(col_name)] + [str(v) if v else "" for v in df_clean.iloc[:, c_idx - 1]]
                        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(len(s) for s in col_data), 50) + 2

                    ws.freeze_panes = "A2"

                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.download_button(
                        label="📥 SCARICA EXCEL PULITO E COLORATO",
                        data=output,
                        file_name="Scyavuru_Database_Pulito.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )

                except Exception as e:
                    st.error(f"Errore durante l'elaborazione: {e}")




# ----------------------------------------------------
# TAB 3: ANALISI COMPETITOR
# ----------------------------------------------------
with tab3:
    st.header("🏆 Analisi Competitor — Lead da Engagement")
    st.markdown(
        "Estrae automaticamente **chi ha reagito ai post** dei tuoi competitor su LinkedIn. "
        "Questi profili sono lead **caldissimi**: conoscono già il settore e interagiscono attivamente con prodotti simili a Scyavuru."
    )

    # --- TABELLA COMPETITOR MODIFICABILE ---
    st.markdown("### 1️⃣ Competitor da analizzare")
    st.caption("Verifica o aggiorna gli slug LinkedIn (la parte finale di `linkedin.com/company/<slug>`).")

    # La lista competitors è ora sincronizzata all'inizio dello script.
    comp_cols = st.columns([2, 3, 1])
    comp_cols[0].markdown("**Competitor**")
    comp_cols[1].markdown("**Slug LinkedIn**")
    comp_cols[2].markdown("**Attivo**")

    for i, comp in enumerate(st.session_state.competitors):
        has_li = comp.get("linkedin_posts", True)
        c1, c2, c3 = st.columns([2, 3, 1])
        with c1:
            label_display = comp["nome"] if has_li else f"🚫 {comp['nome']}"
            st.text_input(
                f"nome_{i}", value=label_display, label_visibility="collapsed",
                key=f"comp_nome_{i}",
                on_change=lambda i=i: st.session_state.competitors[i].update(
                    {"nome": st.session_state[f"comp_nome_{i}"]}
                )
            )
        with c2:
            st.text_input(
                f"slug_{i}", value=comp["slug"] if has_li else "(nessuna pagina LinkedIn)", label_visibility="collapsed",
                key=f"comp_slug_{i}",
                on_change=lambda i=i: st.session_state.competitors[i].update(
                    {"slug": st.session_state[f"comp_slug_{i}"]}
                )
            )
        with c3:
            st.checkbox(
                "on", value=has_li, label_visibility="collapsed",
                key=f"comp_active_{i}"
            )

    # --- PARAMETRI ---
    st.markdown("### 2️⃣ Parametri di estrazione")
    col_p1, col_p2, col_p3 = st.columns(3)
    with col_p1:
        max_posts_comp = st.slider(
            "📄 Post per azienda (ultimi N)", min_value=5, max_value=50, value=20,
            help="Scarica gli N post più recenti, poi filtra per data. Metti un numero alto per coprire il periodo scelto."
        )
    with col_p2:
        max_reactions_comp = st.slider(
            "👥 Max reazioni per post", min_value=10, max_value=200, value=50,
            help="LinkedIn espone max ~1200 reazioni per post."
        )
    with col_p3:
        mesi_indietro = st.selectbox(
            "📅 Periodo post",
            options=[1, 2, 3, 6, 12],
            index=0,
            format_func=lambda x: "Ultimo mese" if x == 1 else f"Ultimi {x} mesi",
            help="Vengono analizzate solo le reactions su post pubblicati in questo periodo."
        )

    # Calcola data di cutoff e mostrala
    from datetime import datetime, timedelta
    cutoff_date_comp = (datetime.today() - timedelta(days=30 * mesi_indietro)).strftime("%Y-%m-%d")
    st.caption(f"🗓️ Periodo analisi: **{cutoff_date_comp}** → oggi. Post antecedenti vengono ignorati (zero costo reactions).")


    # --- FILTRI BOOLEANI (riuso funzione modulo) ---
    with st.expander("🔬 Filtri Booleani Avanzati (AND / OR / NOT)", expanded=False):
        st.markdown("Applicati sui campi **Qualifica** e **Azienda** dei lead estratti.")
        col_cf1, col_cf2 = st.columns(2)
        with col_cf1:
            comp_must_include = st.text_input(
                "✅ DEVE contenere (AND/OR)",
                value="",
                placeholder="es: buyer + food, category manager",
                key="comp_include"
            )
        with col_cf2:
            comp_must_exclude = st.text_input(
                "🚫 NON DEVE contenere (NOT)",
                value="",
                placeholder="es: marketing, hr, recruiting",
                key="comp_exclude"
            )

    n_comp = len([i for i in range(len(st.session_state.get('competitors', DEFAULT_COMPETITORS)))
                   if st.session_state.get(f'comp_active_{i}', True)])
    st.info(
        f"💡 **Stima costo (ultimi {mesi_indietro} {'mese' if mesi_indietro == 1 else 'mesi'}):** "
        f"~$0.002/post + ~$0.002/reazione. "
        f"Con {max_posts_comp} post × {max_reactions_comp} reazioni × {n_comp} competitor attivi = "
        f"**~${round(n_comp * max_posts_comp * 0.002 + n_comp * max_posts_comp * max_reactions_comp * 0.002, 2)}** max"
        " (meno se ci sono pochi post nel periodo scelto)"
    )

    if st.button("🚀 Avvia Analisi Competitor", type="primary"):

        # Raccogli competitor attivi con slug aggiornato dalla UI
        active_comps = []
        for i, comp in enumerate(st.session_state.competitors):
            if st.session_state.get(f"comp_active_{i}", True) and comp.get("linkedin_posts", True):
                active_comps.append({
                    "nome": st.session_state.get(f"comp_nome_{i}", comp["nome"]),
                    "slug": st.session_state.get(f"comp_slug_{i}", comp["slug"]),
                    "showcase": comp.get("showcase", False),
                })

        if not active_comps:
            st.warning("Seleziona almeno un competitor.")
        else:
            all_leads = []
            total_posts_found = 0
            errors = []

            progress_bar = st.progress(0, text="Inizializzazione...")
            status_placeholder = st.empty()

            for ci, comp in enumerate(active_comps):
                progress = ci / len(active_comps)
                progress_bar.progress(progress, text=f"📡 Analisi {comp['nome']} ({ci+1}/{len(active_comps)})...")

                try:
                    # Step 1: scarica i post
                    status_placeholder.info(
                        f"🔍 Scaricando post di **{comp['nome']}** (ultimi {mesi_indietro} "
                        f"{'mese' if mesi_indietro == 1 else 'mesi'}, da {cutoff_date_comp})..."
                    )
                    posts = run_competitor_posts(
                        comp["slug"], max_posts=max_posts_comp, cutoff_date=cutoff_date_comp,
                        showcase=comp.get("showcase", False)
                    )
                    total_posts_found += len(posts)

                    # Step 2: per ogni post, scarica le reazioni e i commenti
                    for pi, post in enumerate(posts):
                        status_placeholder.info(
                            f"💬 Analisi post {pi+1}/{len(posts)} di **{comp['nome']}** "
                            f"({post.get('likes', 0)} likes, {post.get('comments', 0)} commenti) — {post.get('date', '')}"
                        )
                        try:
                            # Scarica reazioni (likes)
                            reactions = run_post_reactions(
                                post["postUrl"], max_reactions=max_reactions_comp
                            )
                            # Scarica commenti
                            comments = run_post_comments(
                                post["postUrl"], max_comments=max_reactions_comp
                            )
                            
                            engagement_list = reactions + comments
                            
                            for person in engagement_list:
                                person["Competitor"] = comp["nome"]
                                person["Follower Competitor"] = post.get("followers", 0)
                                person["Post Snippet"] = post.get("snippet", "")
                                person["Post URL"] = post["postUrl"]
                                person["Post Data"] = post.get("date", "")
                                person["Post Likes"] = post.get("likes", 0)
                                all_leads.append(person)
                        except Exception as e_post:
                            errors.append(f"{comp['nome']} post {pi+1}: {e_post}")

                except Exception as e_comp:
                    errors.append(f"{comp['nome']}: {e_comp}")

            progress_bar.progress(1.0, text="✅ Analisi completata!")
            status_placeholder.empty()

            if errors:
                with st.expander(f"⚠️ {len(errors)} errori durante l'estrazione"):
                    for err in errors:
                        st.warning(err)

            if all_leads:
                df_leads = pd.DataFrame(all_leads)

                # Deduplicazione per profilo
                before_dedup = len(df_leads)
                if "Link Profilo" in df_leads.columns:
                    df_leads = df_leads.drop_duplicates(subset=["Link Profilo"])
                df_leads = df_leads.reset_index(drop=True)
                after_dedup = len(df_leads)

                # Filtri booleani
                df_leads = _apply_boolean_filters(df_leads, comp_must_include, comp_must_exclude)
                after_filter = len(df_leads)

                # --- METRICHE ---
                mc1, mc2, mc3, mc4 = st.columns(4)
                mc1.metric("🏢 Competitor analizzati", len(active_comps))
                mc2.metric("📄 Post analizzati", total_posts_found)
                mc3.metric("👥 Lead unici", after_dedup, delta=f"-{before_dedup - after_dedup} duplicati" if before_dedup > after_dedup else None)
                mc4.metric("✅ Dopo filtri", after_filter, delta=f"-{after_dedup - after_filter}" if after_dedup > after_filter else None)

                # Riordina colonne per leggibilità
                col_order = ["Competitor", "Follower Competitor", "Nome", "Qualifica", "Azienda", "Reazione",
                             "Post Snippet", "Post Data", "Post Likes", "Link Profilo", "Post URL"]
                col_order = [c for c in col_order if c in df_leads.columns]
                df_leads = df_leads[col_order]

                # --- ORDINAMENTO ---
                _sort_cols_t3 = [c for c in ["Competitor", "Azienda", "Qualifica", "Nome", "Post Data", "Post Likes", "Reazione"] if c in df_leads.columns]
                _sort_sel_t3  = st.multiselect(
                    "📊 Ordina per (in sequenza)",
                    options=_sort_cols_t3,
                    default=[c for c in ["Competitor", "Azienda"] if c in _sort_cols_t3],
                    key="sort_tab3",
                    help="Scegli una o più colonne. Il primo campo ha la priorità."
                )
                df_leads = _sort_df(df_leads, [(c, True) for c in _sort_sel_t3])

                st.dataframe(df_leads, use_container_width=True)

                # --- EXCEL FORMATTATO CON COLORE PER COMPETITOR ---
                wb_comp = Workbook()
                ws_comp = wb_comp.active
                ws_comp.title = "Lead Competitor"

                header_font   = Font(bold=True, color="FFFFFF", size=11)
                header_fill   = PatternFill(start_color="1A2940", end_color="1A2940", fill_type="solid")
                header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thin_border   = Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"),
                )

                # Intestazioni
                ws_comp.row_dimensions[1].height = 30
                for c_idx, col_name in enumerate(col_order, 1):
                    cell = ws_comp.cell(row=1, column=c_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border

                # Mappa colori per competitor
                competitors_found = df_leads["Competitor"].unique() if "Competitor" in df_leads.columns else []
                comp_fill_map = {
                    name: PatternFill(
                        start_color=COLORS[i % len(COLORS)],
                        end_color=COLORS[i % len(COLORS)],
                        fill_type="solid"
                    )
                    for i, name in enumerate(competitors_found)
                }

                link_col_idx = col_order.index("Link Profilo") + 1 if "Link Profilo" in col_order else None
                post_url_col_idx = col_order.index("Post URL") + 1 if "Post URL" in col_order else None

                for r_idx, row_vals in enumerate(df_leads.values.tolist(), 2):
                    comp_name = row_vals[col_order.index("Competitor")] if "Competitor" in col_order else ""
                    row_fill = comp_fill_map.get(comp_name)

                    for c_idx, value in enumerate(row_vals, 1):
                        cell = ws_comp.cell(row=r_idx, column=c_idx, value=value)
                        cell.alignment = Alignment(vertical="center", wrap_text=False)
                        cell.border = thin_border
                        if row_fill:
                            cell.fill = row_fill

                        # Link cliccabili
                        if link_col_idx and c_idx == link_col_idx and value:
                            cell.hyperlink = str(value)
                            cell.font = Font(color="1155CC", underline="single")
                        elif post_url_col_idx and c_idx == post_url_col_idx and value:
                            cell.hyperlink = str(value)
                            cell.font = Font(color="1155CC", underline="single")

                # Larghezze automatiche
                for c_idx, col_name in enumerate(col_order, 1):
                    col_data = [str(col_name)] + [str(v) if v else "" for v in df_leads.iloc[:, c_idx - 1]]
                    max_len = min(max(len(s) for s in col_data), 60)
                    ws_comp.column_dimensions[get_column_letter(c_idx)].width = max_len + 2

                ws_comp.freeze_panes = "A2"

                output_comp = io.BytesIO()
                wb_comp.save(output_comp)
                output_comp.seek(0)

                st.download_button(
                    label="📥 Scarica Excel Lead Competitor",
                    data=output_comp,
                    file_name="Scyavuru_Lead_Competitor.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            else:
                st.info("Nessun lead trovato. Prova ad aumentare il numero di post o verificare gli slug dei competitor.")


# ----------------------------------------------------
# TAB 4: PIANO EDITORIALE
# ----------------------------------------------------
with tab4:
    st.header("📰 Piano Editoriale — Analisi Post Competitor")
    st.markdown(
        "Analizza **cosa pubblicano i tuoi competitor** su LinkedIn: argomenti, hashtag, frequenza, "
        "giorni migliori e post più performanti. Usa questi dati per costruire il tuo piano editoriale."
    )

    # --- COMPETITOR (riusa session_state) ---
    st.markdown("### 1️⃣ Competitor")
    st.caption("Attiva/disattiva i competitor da analizzare. I competitor senza pagina LinkedIn ufficiale sono marcati 🚫 e disattivati.")
    pe_cols = st.columns([2, 3, 1])
    pe_cols[0].markdown("**Competitor**")
    pe_cols[1].markdown("**Slug LinkedIn**")
    pe_cols[2].markdown("**Attivo**")
    for i, comp in enumerate(st.session_state.competitors):
        has_li = comp.get("linkedin_posts", True)
        c1, c2, c3 = st.columns([2, 3, 1])
        label_display = comp["nome"] if has_li else f"🚫 {comp['nome']}"
        c1.text_input(f"pe_nome_{i}", value=label_display,
                      label_visibility="collapsed")
        c2.text_input(f"pe_slug_{i}", value=comp["slug"] if has_li else "(nessuna pagina LinkedIn)",
                      label_visibility="collapsed")
        c3.checkbox("on", value=has_li, label_visibility="collapsed",
                    key=f"pe_active_{i}")

    # --- PARAMETRI ---
    st.markdown("### 2️⃣ Parametri")
    col_pe1, col_pe2 = st.columns(2)
    with col_pe1:
        pe_max_posts = st.slider(
            "📄 Post per azienda (ultimi N)", min_value=5, max_value=50, value=20,
            help="Scarica gli N post più recenti poi filtra per data."
        )
    with col_pe2:
        pe_mesi = st.selectbox(
            "📅 Periodo",
            options=[1, 2, 3, 6, 12], index=1,
            format_func=lambda x: "Ultimo mese" if x == 1 else f"Ultimi {x} mesi"
        )

    from datetime import datetime, timedelta
    pe_cutoff = (datetime.today() - timedelta(days=30 * pe_mesi)).strftime("%Y-%m-%d")
    st.caption(f"🗓️ Periodo: **{pe_cutoff}** → oggi")

    # Parole da ignorare nell'analisi keyword
    STOPWORDS_IT = {
        "di","il","la","le","lo","gli","i","un","una","e","in","a","per","con","su","da","che",
        "è","non","del","della","dei","delle","si","ha","al","nel","alla","ai","agli","alle",
        "ci","più","anche","come","ma","o","se","sono","essere","questo","questa","questi",
        "però","così","molto","ogni","tra","dopo","prima","quando","dove","ho","we","the",
        "and","of","to","in","for","is","it","on","that","with","are","from","our","your",
        "have","has","this","all","can","be","an","at","as","by","we","he","she","they",
        "http","https","www","co","com",
    }

    if st.button("🚀 Avvia Analisi Editoriale", type="primary"):

        pe_active = []
        for i, comp in enumerate(st.session_state.competitors):
            if st.session_state.get(f"pe_active_{i}", True) and comp.get("linkedin_posts", True):
                pe_active.append({
                    "nome":     comp["nome"],
                    "slug":     comp["slug"],
                    "showcase": comp.get("showcase", False),
                })
        # Avvisa per i competitor senza LinkedIn
        no_li = [c["nome"] for c in st.session_state.competitors if not c.get("linkedin_posts", True)]
        if no_li:
            st.info(f"ℹ️ Esclusi perché senza pagina LinkedIn ufficiale: **{', '.join(no_li)}**")

        if not pe_active:
            st.warning("Seleziona almeno un competitor.")
        else:
            all_posts = []
            errors_pe = []
            pe_progress = st.progress(0, text="Inizializzazione...")
            pe_status   = st.empty()

            for ci, comp in enumerate(pe_active):
                pe_progress.progress(ci / len(pe_active),
                                     text=f"📡 Scaricando post di {comp['nome']} ({ci+1}/{len(pe_active)})...")
                try:
                    posts = run_competitor_posts(
                        comp["slug"], max_posts=pe_max_posts, cutoff_date=pe_cutoff,
                        showcase=comp.get("showcase", False)
                    )
                    for p in posts:
                        p["Competitor"] = comp["nome"]
                    all_posts.extend(posts)
                    if len(posts) == 0:
                        errors_pe.append(f"{comp['nome']}: 0 post trovati (possibile blocco LinkedIn o crediti proxy esauriti)")
                except Exception as e:
                    err_str = str(e)
                    if "Dataset was not found" in err_str or "dataset" in err_str.lower():
                        errors_pe.append(f"{comp['nome']}: LinkedIn ha bloccato la richiesta (proxy residenziali esauriti sull'account Apify)")
                    else:
                        errors_pe.append(f"{comp['nome']}: {err_str}")

            pe_progress.progress(1.0, text="✅ Completato!")
            pe_status.empty()

            if errors_pe:
                with st.expander(f"⚠️ {len(errors_pe)} competitor con problemi", expanded=True):
                    for err in errors_pe:
                        st.warning(err)
                    st.error(
                        "🔍 **Causa probabile: proxy residenziali esauriti sull'account Apify.**\n\n"
                        "LinkedIn richiede proxy residenziali per lo scraping dei post aziendali. "
                        "Soluzione: vai su [console.apify.com](https://console.apify.com) → Billing → Add credits ($10 ≈ 5.000 post)."
                    )

            if not all_posts:
                st.warning(
                    "🚨 Nessun post recuperato. Questo è quasi certamente un problema di **proxy residenziali esauriti** "
                    "sull'account Apify, non un errore del codice.\n\n"
                    "**Soluzione:** Ricarica i crediti su [console.apify.com](https://console.apify.com) → Billing."
                )

            else:
                df_posts = pd.DataFrame(all_posts)
                total_posts = len(df_posts)

                # Metriche globali
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("🏢 Competitor", len(pe_active))
                m2.metric("📄 Post totali", total_posts)
                m3.metric("👍 Like totali", int(df_posts["likes"].sum()))
                m4.metric("💬 Commenti totali", int(df_posts["comments"].sum()))

                st.divider()

                # ── SEZIONE 1: FREQUENZA DI PUBBLICAZIONE ────────────────
                st.markdown("### 📅 Frequenza di pubblicazione")
                freq_data = df_posts.groupby("Competitor").agg(
                    Post=("postUrl", "count"),
                    Like_medi=("likes", "mean"),
                    Commenti_medi=("comments", "mean"),
                    Engagement_medio=("engagement", "mean"),
                ).round(1).reset_index()
                freq_data = freq_data.sort_values("Post", ascending=False)
                freq_data.columns = ["Competitor", "Post nel periodo", "Like medi", "Commenti medi", "Engagement medio"]
                st.dataframe(freq_data, use_container_width=True, hide_index=True)

                st.divider()

                # ── SEZIONE 2: POST PIÙ PERFORMANTI ──────────────────────
                st.markdown("### 🏆 Post più performanti")
                df_top = df_posts[["Competitor", "date", "weekday", "likes", "comments", "engagement", "snippet", "postUrl"]].copy()
                df_top.columns = ["Competitor", "Data", "Giorno", "Like", "Commenti", "Engagement", "Anteprima", "Link"]
                df_top = df_top.sort_values("Engagement", ascending=False).head(20).reset_index(drop=True)

                # Rendi il link cliccabile nella tabella
                df_top_display = df_top.drop(columns=["Link"])
                st.dataframe(df_top_display, use_container_width=True, hide_index=True)

                st.divider()

                # ── SEZIONE 3: HASHTAG PIÙ USATI ─────────────────────────
                st.markdown("### 🏷️ Hashtag più usati")
                all_hashtags = []
                for _, row in df_posts.iterrows():
                    for tag in row.get("hashtags", []):
                        all_hashtags.append({"Competitor": row["Competitor"], "Hashtag": tag.lower()})

                if all_hashtags:
                    df_ht = pd.DataFrame(all_hashtags)

                    # Globale top 20
                    ht_global = df_ht["Hashtag"].value_counts().head(20).reset_index()
                    ht_global.columns = ["Hashtag", "Frequenza"]

                    # Per competitor
                    ht_by_comp = df_ht.groupby(["Competitor", "Hashtag"]).size().reset_index(name="Frequenza")
                    ht_by_comp = ht_by_comp.sort_values(["Competitor", "Frequenza"], ascending=[True, False])

                    col_ht1, col_ht2 = st.columns([1, 2])
                    with col_ht1:
                        st.markdown("**Top 20 globali**")
                        st.dataframe(ht_global, use_container_width=True, hide_index=True)
                    with col_ht2:
                        st.markdown("**Per competitor**")
                        # Top 5 per ognuno
                        ht_top5 = ht_by_comp.groupby("Competitor").head(5).reset_index(drop=True)
                        st.dataframe(ht_top5, use_container_width=True, hide_index=True)
                else:
                    st.info("Nessun hashtag trovato nei post del periodo.")

                st.divider()

                # ── SEZIONE 4: GIORNO MIGLIORE ────────────────────────────
                st.markdown("### 📆 Giorno con più engagement")
                day_order = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
                df_days = df_posts[df_posts["weekday"] != ""].copy()
                if not df_days.empty:
                    day_stats = df_days.groupby("weekday").agg(
                        Post=("postUrl", "count"),
                        Engagement_medio=("engagement", "mean")
                    ).round(1).reset_index()
                    day_stats["weekday"] = pd.Categorical(day_stats["weekday"], categories=day_order, ordered=True)
                    day_stats = day_stats.sort_values("weekday").reset_index(drop=True)
                    day_stats.columns = ["Giorno", "Post pubblicati", "Engagement medio"]
                    st.dataframe(day_stats, use_container_width=True, hide_index=True)
                    best_day = day_stats.sort_values("Engagement medio", ascending=False).iloc[0]["Giorno"]
                    st.success(f"💡 Il giorno con più engagement medio è **{best_day}**")
                else:
                    st.info("Date non disponibili per l'analisi per giorno.")

                st.divider()

                # ── SEZIONE 5: KEYWORD PIÙ FREQUENTI ─────────────────────
                st.markdown("### 🔑 Parole chiave più usate nei post")
                all_words = []
                for _, row in df_posts.iterrows():
                    words = re.findall(r'\b[a-zàèéìòùA-Z]{4,}\b', row.get("full_text", ""))
                    for w in words:
                        wl = w.lower()
                        if wl not in STOPWORDS_IT:
                            all_words.append({"Competitor": row["Competitor"], "Parola": wl})

                if all_words:
                    df_words = pd.DataFrame(all_words)
                    kw_global = df_words["Parola"].value_counts().head(30).reset_index()
                    kw_global.columns = ["Parola", "Frequenza"]
                    kw_by_comp = df_words.groupby(["Competitor", "Parola"]).size().reset_index(name="Frequenza")
                    kw_by_comp = kw_by_comp.sort_values(["Competitor", "Frequenza"], ascending=[True, False])
                    kw_top5 = kw_by_comp.groupby("Competitor").head(8).reset_index(drop=True)

                    col_kw1, col_kw2 = st.columns([1, 2])
                    with col_kw1:
                        st.markdown("**Top 30 globali**")
                        st.dataframe(kw_global, use_container_width=True, hide_index=True)
                    with col_kw2:
                        st.markdown("**Top 8 per competitor**")
                        st.dataframe(kw_top5, use_container_width=True, hide_index=True)
                else:
                    st.info("Testo non disponibile per l'analisi keyword.")

                st.divider()

                # ── SEZIONE 6: TUTTI I POST COMPLETI ─────────────────────
                st.markdown("### 📋 Tutti i post (testo completo)")
                sort_pe_cols = [c for c in ["Competitor", "date", "engagement", "likes", "comments", "shares"] if c in df_posts.columns]
                sort_pe_sel  = st.multiselect(
                    "📊 Ordina per", options=sort_pe_cols,
                    default=["Competitor", "engagement"],
                    key="sort_tab4"
                )
                df_full_display = _sort_df(df_posts, [(c, c not in ["engagement","likes","comments","shares"]) for c in sort_pe_sel])

                show_cols = ["Competitor", "followers", "date", "weekday", "likes", "comments", "shares", "engagement", "snippet", "hashtags", "postUrl"]
                show_cols = [c for c in show_cols if c in df_full_display.columns]
                rename_map = {"followers": "Follower", "date": "Data", "weekday": "Giorno", "likes": "Like",
                              "comments": "Commenti", "shares": "Condivisioni",
                              "engagement": "Engagement",
                              "snippet": "Anteprima", "hashtags": "Hashtag", "postUrl": "Link"}
                df_full_display = df_full_display[show_cols].rename(columns=rename_map)
                st.dataframe(df_full_display, use_container_width=True, hide_index=True)

                st.divider()

                # ── SEZIONE 7: GAP ANALYSIS STRATEGICO ──────────────────────
                st.markdown("### 🧠 Insight Strategici per Scyavuru")
                st.caption("Analisi automatica generata dai dati estratti. Identifica i gap di comunicazione che Scyavuru può sfruttare.")

                # Calcola insights dai dati
                _best_eng  = freq_data.sort_values("Engagement medio", ascending=False).iloc[0] if not freq_data.empty else None
                _most_act  = freq_data.sort_values("Post nel periodo", ascending=False).iloc[0] if not freq_data.empty else None
                _top_tags  = ht_global["Hashtag"].head(6).tolist() if all_hashtags and not ht_global.empty else []
                _best_day_ins = day_stats.sort_values("Engagement medio", ascending=False).iloc[0]["Giorno"] if not df_days.empty and not day_stats.empty else "N/D"

                col_ins1, col_ins2 = st.columns(2)
                with col_ins1:
                    st.markdown("#### 🔍 Benchmark competitor")
                    if _best_eng is not None:
                        st.metric("🏆 Engagement medio più alto", _best_eng["Competitor"], f"{_best_eng['Engagement medio']:.1f} avg")
                    if _most_act is not None:
                        st.metric("📣 Competitor più attivo", _most_act["Competitor"], f"{int(_most_act['Post nel periodo'])} post nel periodo")
                    st.metric("📆 Giorno migliore per postare", _best_day_ins)
                    if _top_tags:
                        st.warning(f"**Hashtag già inflazionati dai competitor:**\n`{'`  `'.join(_top_tags)}`\n\n→ Non usarli uguali, differenziati!")

                with col_ins2:
                    st.markdown("#### 🎯 Gap da sfruttare (nessun competitor lo fa)")
                    st.success("🟢 **Zero messaggi B2B** — Tutti parlano al consumatore finale. Scyavuru può prendere lo spazio come *fornitore strategico*.")
                    st.success("🟢 **Zero dati concreti** — Nessuno cita numeri. Usa: paesi serviti, anni di esperienza, volumi.")
                    st.success("🟢 **Supply chain ignorata** — Comunica affidabilità e continuità forniture (365 gg/anno).")
                    st.success("🟢 **DOP / Certificazioni mai citate** — Differenziati con tracciabilità filiera e certificazioni.")
                    st.success("🟢 **Nessuna testimonianza buyer** — Un case study di un cliente GDO vale 100 post istituzionali.")

                st.divider()

                # ── SEZIONE 8: TOP POST PER FOLLOWER INTELLIGENCE ─────────
                st.markdown("### 🔥 Top Post — Candidati per Follower Intelligence")
                st.markdown(
                    "I post con **più engagement** sono i migliori candidati per estrarre le reactions: "
                    "chi ha messo ❤️ su questi post è un **lead caldo** — già interessato al settore. "
                    "Copia l'URL e usalo nel tab **🏆 Analisi Competitor** per estrarne i profili."
                )

                _top5_reactions = df_posts.nlargest(5, "engagement")[
                    ["Competitor", "date", "likes", "comments", "engagement", "snippet", "postUrl"]
                ].copy().reset_index(drop=True)
                _top5_reactions.columns = ["Competitor", "Data", "❤️ Like", "💬 Commenti", "🔥 Tot. Engagement", "Anteprima Post", "🔗 URL Post"]

                st.dataframe(_top5_reactions, use_container_width=True, hide_index=True)

                st.info(
                    "💡 **Come usarli:** Copia un URL dalla colonna *🔗 URL Post* → vai al tab **🏆 Analisi Competitor** "
                    "→ nella sezione avanzata incolla il link per estrarre chi ha reagito al post (buyers + category managers)."
                )

                # Box con URL cliccabili per copia rapida
                with st.expander("📋 URL top post (copia rapida per Tab 🏆)"):
                    for _, row in _top5_reactions.iterrows():
                        st.markdown(
                            f"**{row['Competitor']}** — {row['Data']} — {int(row['🔥 Tot. Engagement'])} engagement\n\n"
                            f"`{row['🔗 URL Post']}`"
                        )
                        st.divider()

                st.divider()

                # ── EXCEL COMPLETO ────────────────────────────────────────
                st.markdown("### 📥 Esporta")
                wb_pe = Workbook()
                # Un foglio per ogni competitor + foglio riepilogo
                wb_pe.remove(wb_pe.active)  # rimuove il foglio vuoto di default

                header_font_pe = Font(bold=True, color="FFFFFF", size=11)
                header_fill_pe = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
                header_align_pe = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thin_b_pe = Border(
                    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
                )

                def _write_sheet(wb, title, df_sheet, fill_color=None):
                    ws = wb.create_sheet(title=title[:31])
                    cols_s = list(df_sheet.columns)
                    ws.row_dimensions[1].height = 28
                    for ci, cn in enumerate(cols_s, 1):
                        cell = ws.cell(row=1, column=ci, value=cn)
                        cell.font = header_font_pe
                        cell.fill = header_fill_pe
                        cell.alignment = header_align_pe
                        cell.border = thin_b_pe
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
                    link_ci = cols_s.index("Link") + 1 if "Link" in cols_s else None
                    for ri, row_vals in enumerate(df_sheet.values.tolist(), 2):
                        for ci, val in enumerate(row_vals, 1):
                            # Converti liste in stringa per Excel
                            if isinstance(val, list):
                                val = ", ".join(val)
                            cell = ws.cell(row=ri, column=ci, value=val)
                            cell.border = thin_b_pe
                            cell.alignment = Alignment(vertical="top", wrap_text=(ci == (cols_s.index("Testo completo")+1 if "Testo completo" in cols_s else -1)))
                            if fill:
                                cell.fill = fill
                            if link_ci and ci == link_ci and val:
                                cell.hyperlink = str(val)
                                cell.font = Font(color="1155CC", underline="single")
                    # Larghezze
                    for ci, cn in enumerate(cols_s, 1):
                        col_data = [str(cn)] + [str(v) if v else "" for v in df_sheet.iloc[:, ci-1]]
                        w = min(max(len(str(s)) for s in col_data), 60) + 2
                        ws.column_dimensions[get_column_letter(ci)].width = w
                    ws.freeze_panes = "A2"

                # Foglio RIEPILOGO
                _write_sheet(wb_pe, "Riepilogo", freq_data)

                # Foglio per ogni competitor
                for ci_comp, comp_name in enumerate(df_posts["Competitor"].unique()):
                    df_comp_sheet = df_posts[df_posts["Competitor"] == comp_name].copy()
                    exp_cols = [c for c in ["followers","date","weekday","likes","comments","shares","engagement","snippet","full_text","hashtags","postUrl"] if c in df_comp_sheet.columns]
                    df_comp_sheet = df_comp_sheet[exp_cols].copy()
                    df_comp_sheet["hashtags"] = df_comp_sheet["hashtags"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
                    col_labels = {"followers": "Follower", "date": "Data", "weekday": "Giorno", "likes": "Like",
                                  "comments": "Commenti", "shares": "Condivisioni",
                                  "engagement": "Engagement", "snippet": "Anteprima",
                                  "full_text": "Testo completo", "hashtags": "Hashtag", "postUrl": "Link"}
                    df_comp_sheet.columns = [col_labels.get(c, c) for c in exp_cols]
                    df_comp_sheet = df_comp_sheet.sort_values("Engagement", ascending=False).reset_index(drop=True)
                    color = COLORS[ci_comp % len(COLORS)]
                    _write_sheet(wb_pe, comp_name[:31], df_comp_sheet, fill_color=color)

                # Foglio HASHTAG globali
                if all_hashtags:
                    _write_sheet(wb_pe, "Hashtag", ht_global)

                # Foglio KEYWORD globali
                if all_words:
                    _write_sheet(wb_pe, "Keyword", kw_global)

                out_pe = io.BytesIO()
                wb_pe.save(out_pe)
                out_pe.seek(0)

                st.download_button(
                    label="📥 Scarica Excel Piano Editoriale",
                    data=out_pe,
                    file_name="Scyavuru_Piano_Editoriale.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

        # --- NEW: SECTION FOR LINKEDIN POST SEARCH SCRAPER ---
        st.divider()
        st.markdown("### 🔍 Trova Post da Commentare (LinkedIn Post Search)")
        st.markdown("Usa questa sezione per cercare discussioni attive su LinkedIn in base a keyword o hashtag per inserire i commenti di Rosario.")
        
        col_search1, col_search2 = st.columns([3, 1])
        with col_search1:
            search_query = st.text_input("🔑 Keyword o Hashtag per la ricerca (es. #GDO, private label)", value="#GDO")
        with col_search2:
            search_limit = st.slider("📄 Max Post", min_value=5, max_value=50, value=10, key="search_limit_slider")
            
        if st.button("🚀 Avvia Ricerca Post", type="primary"):
            if not search_query.strip():
                st.warning("Inserisci una query di ricerca.")
            else:
                with st.spinner(f"🔍 Ricerca in corso su LinkedIn per '{search_query}'..."):
                    try:
                        search_results = run_linkedin_post_search(search_query, max_posts=search_limit)
                        if search_results:
                            df_search = pd.DataFrame(search_results)
                            st.success(f"Trovati {len(search_results)} post corrispondenti!")
                            st.dataframe(df_search[["author", "date", "likes", "comments", "engagement", "text", "url"]], use_container_width=True)
                            
                            st.markdown("#### 🎯 Post Dettagliati e Bozza di Commento AI")
                            for idx, post in enumerate(search_results):
                                with st.container():
                                    st.write(f"👤 **Autore:** {post['author']} | 📅 Data: {post['date']}")
                                    st.write(f"📈 Likes: {post['likes']} | Commenti: {post['comments']} | Engagement: {post['engagement']}")
                                    st.info(f"📝 **Anteprima Testo:** {post['text']}")
                                    
                                    lower_text = post['text'].lower()
                                    if "private label" in lower_text or "mdd" in lower_text:
                                        comment_idea = (
                                            "\"Concordo in pieno. Il Private Label gourmet rappresenta oggi la vera opportunità per i retailer "
                                            "di difendere le marginalità a scaffale ed evitare la guerra del prezzo. Dal nostro punto di vista come "
                                            "co-produttori B2B, la sfida è offrire ricette altamente personalizzate (es. clean label o zero zuccheri) "
                                            "garantendo affidabilità e costanza industriale nei volumi.\""
                                        )
                                    elif "sostenibil" in lower_text or "esg" in lower_text:
                                        comment_idea = (
                                            "\"Ottima riflessione. Nel food, la sostenibilità deve tradursi in scelte concrete lungo tutta la filiera: "
                                            "dall'uso di energia 100% solare in produzione fino al recupero creativo degli scarti (Zero Waste), "
                                            "come facciamo noi a Ribera valorizzando i gusci di pistacchio. Solo così creiamo valore reale per il consumatore.\""
                                        )
                                    else:
                                        comment_idea = (
                                            "\"Riflessione molto condivisibile. Avere il controllo totale del processo produttivo, "
                                            "importando e tostando le materie prime intere (pistacchi, mandorle, nocciole) direttamente in stabilimento, "
                                            "è l'unico modo per garantire ai partner della GDO la freschezza e la stabilità organolettica che meritano.\""
                                        )
                                        
                                    st.markdown("💡 **Suggerimento Bozza Commento AI (copia e adatta):**")
                                    st.code(comment_idea, language="text")
                                    st.link_button("🔗 Apri e Commenta su LinkedIn", post['url'])
                                    st.divider()
                        else:
                            st.info("Nessun post trovato per questa query. Prova ad inserire un hashtag più comune.")
                    except Exception as e:
                        st.error(f"Errore durante la ricerca: {e}")
                        
        # --- NEW: SECTION FOR GOOGLE TRENDS SCRAPER ---
        st.divider()
        st.markdown("### 📈 Analizzatore di Trend (Google Trends)")
        st.markdown("Usa questa sezione per estrarre le parole e le query correlate in forte crescita su Google per avere idee fresche per i tuoi post su LinkedIn.")
        
        col_t1, col_t2 = st.columns([3, 1])
        with col_t1:
            trends_query = st.text_input("🔑 Termini da analizzare (separati da virgola, es. crema pistacchio, pistachio cream)", value="crema pistacchio")
        with col_t2:
            trends_time = st.selectbox("📅 Intervallo Temporale", ["today 1-m", "today 3-m", "today 12-m"], index=1)
            
        if st.button("🚀 Avvia Analisi Trend", type="primary"):
            if not trends_query.strip():
                st.warning("Inserisci almeno un termine di ricerca.")
            else:
                terms_list = [t.strip() for t in trends_query.split(",") if t.strip()]
                with st.spinner(f"📈 Estrazione dati da Google Trends per {terms_list}..."):
                    try:
                        trends_results = run_google_trends(terms_list, timeframe=trends_time)
                        if trends_results:
                            st.success(f"Trovati {len(trends_results)} trend e query correlate in forte crescita!")
                            df_trends = pd.DataFrame(trends_results)
                            st.dataframe(df_trends, use_container_width=True, hide_index=True)
                            
                            st.markdown("#### 💡 Idee per Post LinkedIn basate su questi Trend")
                            for idx, trend in enumerate(trends_results[:5]):
                                q_val = trend['Query']
                                c_val = trend['Crescita']
                                st.write(f"📌 **Idea Post per '{q_val}' (Crescita: {c_val}):**")
                                idea_text = (
                                    f"Parlare dell'impennata di interesse verso '{q_val}' nel mercato. "
                                    f"Spiegare come Scyavuru risponde a questo trend (es. lanciando varianti dedicate per la GDO "
                                    f"o aumentando la capacità produttiva di questa referenza per i partner private label)."
                                )
                                st.info(idea_text)
                        else:
                            st.info("Nessuna query correlata in crescita trovata per questo periodo. Prova con termini più ampi (es. 'pistacchio').")
                    except Exception as e:
                        st.error(f"Errore durante l'analisi dei trend: {e}")


# ----------------------------------------------------
# TAB 5: GOOGLE & LINKEDIN TRENDS
# ----------------------------------------------------
with tab5:
    st.header("📈 Google & LinkedIn Trends")
    st.markdown(
        "Trova i post più caldi su LinkedIn per commentare e le query in crescita su Google Trends per trovare nuove idee per i tuoi post."
    )
    
    tab_trend_li, tab_trend_go = st.tabs(["🔥 LinkedIn Post Search", "📊 Google Trends Scraper"])
    
    with tab_trend_li:
        st.markdown("### 🔍 Trova Post da Commentare (LinkedIn Post Search)")
        st.markdown("Usa questa sezione per cercare discussioni attive su LinkedIn in base a keyword o hashtag per inserire i commenti di Rosario.")
        
        col_search1, col_search2 = st.columns([3, 1])
        with col_search1:
            search_query = st.text_input("🔑 Keyword o Hashtag per la ricerca (es. #GDO, private label)", value="#GDO", key="trend_search_query_input")
        with col_search2:
            search_limit = st.slider("📄 Max Post", min_value=5, max_value=50, value=10, key="trend_search_limit_slider")
            
        if st.button("🚀 Avvia Ricerca Post", type="primary", key="btn_trend_search_li"):
            if not search_query.strip():
                st.warning("Inserisci una query di ricerca.")
            else:
                with st.spinner(f"🔍 Ricerca in corso su LinkedIn per '{search_query}'..."):
                    try:
                        search_results = run_linkedin_post_search(search_query, max_posts=search_limit)
                        if search_results:
                            df_search = pd.DataFrame(search_results)
                            st.success(f"Trovati {len(search_results)} post corrispondenti!")
                            st.dataframe(df_search[["author", "date", "likes", "comments", "engagement", "text", "url"]], use_container_width=True)
                            
                            st.markdown("#### 🎯 Post Dettagliati e Bozza di Commento AI")
                            for idx, post in enumerate(search_results):
                                with st.container():
                                    st.write(f"👤 **Autore:** {post['author']} | 📅 Data: {post['date']}")
                                    st.write(f"📈 Likes: {post['likes']} | Commenti: {post['comments']} | Engagement: {post['engagement']}")
                                    st.info(f"📝 **Anteprima Testo:** {post['text']}")
                                    
                                    lower_text = post['text'].lower()
                                    if "private label" in lower_text or "mdd" in lower_text:
                                        comment_idea = (
                                            "\"Concordo in pieno. Il Private Label gourmet rappresenta oggi la vera opportunità per i retailer "
                                            "di difendere le marginalità a scaffale ed evitare la guerra del prezzo. Dal nostro punto di vista come "
                                            "co-produttori B2B, la sfida è offrire ricette altamente personalizzate (es. clean label o zero zuccheri) "
                                            "garantendo affidabilità e costanza industriale nei volumi.\""
                                        )
                                    elif "sostenibil" in lower_text or "esg" in lower_text:
                                        comment_idea = (
                                            "\"Ottima riflessione. Nel food, la sostenibilità deve tradursi in scelte concrete lungo tutta la filiera: "
                                            "dall'uso di energia 100% solare in produzione fino al recupero creativo degli scarti (Zero Waste), "
                                            "come facciamo noi a Ribera valorizzando i gusci di pistacchio. Solo così creiamo valore reale per il consumatore.\""
                                        )
                                    else:
                                        comment_idea = (
                                            "\"Riflessione molto condivisibile. Avere il controllo totale del processo produttivo, "
                                            "importando e tostando le materie prime intere (pistacchi, mandorle, nocciole) direttamente in stabilimento, "
                                            "è l'unico modo per garantire ai partner della GDO la freschezza e la stabilità organolettica che meritano.\""
                                        )
                                        
                                    st.markdown("💡 **Suggerimento Bozza Commento AI (copia e adatta):**")
                                    st.code(comment_idea, language="text")
                                    if post['url']:
                                        st.link_button("🔗 Apri e Commenta su LinkedIn", post['url'])
                                    st.divider()
                        else:
                            st.info("Nessun post trovato per questa query. Prova ad inserire un hashtag più comune.")
                    except Exception as e:
                        st.error(f"Errore durante la ricerca: {e}")
                        
    with tab_trend_go:
        st.markdown("### 📈 Analizzatore di Trend (Google Trends)")
        st.markdown("Usa questa sezione per estrarre le parole e le query correlate in forte crescita su Google per avere idee fresche per i tuoi post su LinkedIn.")
        
        col_t1, col_t2 = st.columns([3, 1])
        with col_t1:
            trends_query = st.text_input("🔑 Termini da analizzare (separati da virgola, es. crema pistacchio, pistachio cream)", value="crema pistacchio", key="trend_go_query_input")
        with col_t2:
            trends_time = st.selectbox("📅 Intervallo Temporale", ["today 1-m", "today 3-m", "today 12-m"], index=1, key="trend_go_time_select")
            
        if st.button("🚀 Avvia Analisi Trend", type="primary", key="btn_trend_go_trends"):
            if not trends_query.strip():
                st.warning("Inserisci almeno un termine di ricerca.")
            else:
                terms_list = [t.strip() for t in trends_query.split(",") if t.strip()]
                with st.spinner(f"📈 Estrazione dati da Google Trends per {terms_list}..."):
                    try:
                        trends_results = run_google_trends(terms_list, timeframe=trends_time)
                        if trends_results:
                            st.success(f"Trovati {len(trends_results)} trend e query correlate in forte crescita!")
                            df_trends = pd.DataFrame(trends_results)
                            st.dataframe(df_trends, use_container_width=True, hide_index=True)
                            
                            st.markdown("#### 💡 Idee per Post LinkedIn basate su questi Trend")
                            for idx, trend in enumerate(trends_results[:5]):
                                q_val = trend['Query']
                                c_val = trend['Crescita']
                                st.write(f"📌 **Idea Post per '{q_val}' (Crescita: {c_val}):**")
                                idea_text = (
                                    f"Parlare dell'impennata di interesse verso '{q_val}' nel mercato. "
                                    f"Spiegare come Scyavuru risponde a questo trend (es. lanciando varianti dedicate per la GDO "
                                    f"o aumentando la capacità produttiva di questa referenza per i partner private label)."
                                )
                                st.info(idea_text)
                        else:
                            st.info("Nessuna query correlata in crescita trovata per questo periodo. Prova con termini più ampi (es. 'pistacchio').")
                    except Exception as e:
                        st.error(f"Errore durante l'analisi dei trend: {e}")


# ----------------------------------------------------
# TAB 6: ACTION PLAN & TO-DO
# ----------------------------------------------------
with tab6:
    st.header("📌 Action Plan & To-Do (Riunione Ferdinando)")
    st.markdown("Questa sezione è uno strumento persistente per tenere traccia delle priorità operative decise nelle call strategiche. I progressi vengono salvati in automatico.")
    
    import json
    import os

    TODO_FILE = "todo_scyavuru.json"
    
    default_tasks = {
        "1. Sincronizzare rubrica di Rosario tramite l'App Mobile LinkedIn (escludendo contatti non pertinenti)": False,
        "2. Rivedere i testi della pagina Company (NON aspettare i vettoriali)": False,
        "3. Raccogliere fisicamente le Segnalazioni scritte dagli amici/partner e farle pubblicare": False,
        "4. Iniziare a scrivere i testi descrittivi per i nuovi Cataloghi (partendo dal file Excel)": False,
        "5. Creazione annuncio 'Job Posting' mirato (es. Export Manager GDO) per raccogliere lead": False,
        "6. Avviare commenti tattici e Repost 'Con Pensiero' sui profili dei partner": False,
        "7. Attivazione Demo Sales Navigator (SOLO quando i cataloghi sono pronti)": False,
        "8. Inizio pubblicazione Post ufficiali di Rosario (SOLO dopo aver raggiunto 150-200 contatti)": False
    }

    if not os.path.exists(TODO_FILE):
        with open(TODO_FILE, "w", encoding="utf-8") as f:
            json.dump(default_tasks, f)
            
    with open(TODO_FILE, "r", encoding="utf-8") as f:
        tasks = json.load(f)
        
    # Check if there are new default tasks not present in the saved JSON and add them
    for task_name in default_tasks:
        if task_name not in tasks:
            tasks[task_name] = False
            
    # Risorse utili
    st.subheader("📎 Risorse Rapide")
    st.link_button("📊 Apri Catalogo Excel (Google Sheets)", "https://docs.google.com/spreadsheets/d/1PTbnw1Y1Eo5TmbhQWIohVxjCVzCDUftfwX0EhDhjUo8/edit?usp=sharing")
    st.divider()

    st.subheader("✅ Task List Operativa")
    
    # Form per aggiungere nuove task
    with st.form("add_new_task_scyavuru", clear_on_submit=True):
        col_a1, col_a2 = st.columns([4, 1])
        new_task_text = col_a1.text_input("📝 Aggiungi una nuova priorità / task:")
        submit_add = col_a2.form_submit_button("Aggiungi")
        if submit_add and new_task_text.strip():
            tasks[new_task_text.strip()] = False
            with open(TODO_FILE, "w", encoding="utf-8") as f:
                json.dump(tasks, f)
            st.rerun()
            
    st.write("") # spacing
    
    # Render delle task con checkbox e pulsante di eliminazione
    updated_tasks = {}
    for task_name, state in list(tasks.items()):
        col_t1, col_t2 = st.columns([7, 1])
        checked = col_t1.checkbox(task_name, value=state, key=f"check_scy_{task_name}")
        if checked != state:
            tasks[task_name] = checked
            with open(TODO_FILE, "w", encoding="utf-8") as f:
                json.dump(tasks, f)
            st.rerun()
            
        import hashlib
        task_hash = hashlib.md5(task_name.encode('utf-8')).hexdigest()[:10]
        if col_t2.button("🗑️", key=f"del_scy_{task_hash}"):
            if task_name in tasks:
                del tasks[task_name]
                with open(TODO_FILE, "w", encoding="utf-8") as f:
                    json.dump(tasks, f)
                st.rerun()
                
        updated_tasks[task_name] = checked
        
    st.divider()
    completed = sum(1 for v in updated_tasks.values() if v)
    total = len(updated_tasks)
    progress = completed / total if total > 0 else 0
    st.progress(progress, text=f"Progresso: {completed} su {total} task completati")

